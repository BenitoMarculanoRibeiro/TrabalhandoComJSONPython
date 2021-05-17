import json 
from openpyxl import Workbook


def ler_json(arq_json):
    with open(arq_json, 'r', encoding='utf8') as f:
        return json.load(f)


data = ler_json('Java\pedido1.json')

nTempo = list(data["Pedido 1"][0].keys())[0]
Tempo = list(data["Pedido 1"][0].values())[0]
nFitness = list(data["Pedido 1"][1].keys())[0]
Fitness = list(data["Pedido 1"][1].values())[0]
nPopulacao = list(data["Pedido 1"][2].keys())[0]
Populacao = list(data["Pedido 1"][2].values())[0]
nSelecao = list(data["Pedido 1"][3].keys())[0]
Selecao = list(data["Pedido 1"][3].values())[0]
nCruzamento = list(data["Pedido 1"][4].keys())[0]
Cruzamento = list(data["Pedido 1"][4].values())[0]
nMutacao = list(data["Pedido 1"][5].keys())[0]
Mutacao = list(data["Pedido 1"][5].values())[0]
nInsercao = list(data["Pedido 1"][6].keys())[0]
Insercao = list(data["Pedido 1"][6].values())[0]


wb = Workbook()

# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
ws['A1'] = nTempo
for i, statN in enumerate(Tempo):
    ws.cell(row=i+2, column=1).value = statN
    
ws['B1'] = nFitness
for i, statN in enumerate(Fitness):
    ws.cell(row=i+2, column=2).value = statN

ws['C1'] = nPopulacao
for i, statN in enumerate(Populacao):
    ws.cell(row=i+2, column=3).value = statN
    
ws['D1'] = nSelecao
for i, statN in enumerate(Selecao):
    ws.cell(row=i+2, column=4).value = statN
    
ws['E1'] = nCruzamento
for i, statN in enumerate(Cruzamento):
    ws.cell(row=i+2, column=5).value = statN

ws['F1'] = nMutacao
for i, statN in enumerate(Mutacao):
    ws.cell(row=i+2, column=6).value = statN

ws['G1'] = nInsercao
for i, statN in enumerate(Insercao):
    ws.cell(row=i+2, column=7).value = statN

# Save the file
wb.save("Excel\Java\pedido1.xlsx")